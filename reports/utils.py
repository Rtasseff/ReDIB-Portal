"""
Utility functions for report generation.
Based on design document section 9 - Reporting and Statistics.
"""

from openpyxl import Workbook
from django.db.models import Sum, Avg, Count
from applications.models import Application
from access.models import Publication


def generate_call_report_excel(call):
    """
    Generate Excel workbook for call summary report.

    Creates a 3-sheet workbook:
    - Summary: Call info and application statistics
    - Applications: Detailed application list
    - Equipment Utilization: Equipment usage metrics

    Args:
        call: Call object to generate report for

    Returns:
        openpyxl.Workbook object
    """
    wb = Workbook()

    # Sheet 1: Summary
    ws1 = wb.active
    ws1.title = "Summary"
    ws1.append(['Call Report: ' + call.code])
    ws1.append([])
    ws1.append(['Call Information'])
    ws1.append(['Code', call.code])
    ws1.append(['Title', call.title])
    ws1.append(['Submission Period', f"{call.submission_start.date()} to {call.submission_end.date()}"])
    ws1.append(['Evaluation Deadline', call.evaluation_deadline.date() if call.evaluation_deadline else 'N/A'])
    ws1.append([])

    # Application statistics
    apps = Application.objects.filter(call=call)
    publications = Publication.objects.filter(application__call=call)

    ws1.append(['Application Statistics'])
    ws1.append(['Total Applications', apps.count()])
    ws1.append(['Accepted', apps.filter(resolution='accepted').count()])
    ws1.append(['Rejected', apps.filter(resolution='rejected').count()])
    ws1.append(['Pending', apps.filter(resolution='pending').count()])

    # Calculate average score (handle None values)
    avg_score = apps.aggregate(avg=Avg('final_score'))['avg']
    ws1.append(['Average Evaluation Score', f"{float(avg_score):.2f}" if avg_score else 'N/A'])

    # Calculate acceptance rate
    total_apps = apps.count()
    accepted_apps = apps.filter(resolution='accepted').count()
    acceptance_rate = (accepted_apps / total_apps * 100) if total_apps > 0 else 0
    ws1.append(['Acceptance Rate', f"{acceptance_rate:.1f}%"])

    ws1.append([])
    ws1.append(['Publication Tracking'])
    ws1.append(['Publications Reported', publications.count()])
    ws1.append(['Publications with Acknowledgment', publications.filter(redib_acknowledged=True).count()])

    # Sheet 2: Applications List
    ws2 = wb.create_sheet("Applications")
    ws2.append([
        'Code', 'Applicant', 'Institution', 'Status', 'Final Score',
        'Resolution', 'Hours Requested'
    ])

    for app in apps.select_related('applicant').prefetch_related('requested_access'):
        # Calculate total hours from requested_access
        total_requested = app.requested_access.aggregate(
            total=Sum('hours_requested')
        )['total'] or 0

        ws2.append([
            app.code or 'DRAFT',
            app.applicant_name or app.applicant.get_full_name(),
            app.applicant_entity or '',
            app.get_status_display(),
            float(app.final_score) if app.final_score else None,
            app.get_resolution_display() if app.resolution else '',
            float(total_requested),
        ])

    # Sheet 3: Equipment Summary
    ws3 = wb.create_sheet("Equipment Summary")
    ws3.append([
        'Equipment', 'Node', 'Total Approved Hours'
    ])

    for allocation in call.equipment_allocations.select_related('equipment__node').all():
        total_approved = allocation.total_approved_hours  # Sum of hours_requested from accepted apps

        ws3.append([
            allocation.equipment.name,
            allocation.equipment.node.code,
            float(total_approved),
        ])

    return wb
